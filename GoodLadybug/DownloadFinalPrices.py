import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from webdriver_manager.chrome import ChromeDriverManager

# The URL of the auction page
AUCTION_URL = input("Enter the URL name: ")

# Path to your Excel file on your local machine
EXCEL_PATH = input("Enter the path to the Excel file: ")
SHEET_NAME = input("Enter the sheet name: ")

def setup_selenium_driver():
    # Set up the Selenium Chrome driver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    return driver

def read_lot_and_names_from_excel(excel_path, sheet_name):
    # Load the workbook and the sheet
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Read Lot numbers and Names from Excel and store them in a dictionary
    lot_to_name = {}
    for row in sheet.iter_rows(min_row=10, max_col=11, values_only=True):
        lot_number = row[10]  # Assuming lot numbers are in column K
        name = row[1]  # Assuming names are in column B
        lot_to_name[lot_number] = name
    return lot_to_name

def get_lot_details_from_listing(listing):
    # Find the lot name using the provided HTML structure
    lot_name_element = listing.find('p', class_='catalogList-desc my-1')
    lot_name = lot_name_element.get_text(strip=True) if lot_name_element else None

    # Find the lot number from the <h4> tag with the auction-Itemlist-Title class
    lot_number_element = listing.find('h4', class_='auction-Itemlist-Title').find('a')
    lot_number = lot_number_element.get_text(strip=True) if lot_number_element else None

    return lot_number, lot_name

def get_price_from_listing(listing):
    # Extract the price and check for a multiplier
    price_text = listing.find('span', class_='font-1rem').get_text(strip=True)
    price_pattern = re.compile(r'(\d+\.?\d*)\s*(?:\( x (\d+) \))?')
    match = price_pattern.search(price_text)
    if match:
        price_value, multiplier = match.groups(default="1")
        return float(price_value.replace(',', '')) * int(multiplier)
    return None

def scrape_and_write_prices(driver, excel_path, sheet_name, lot_to_name):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Navigate through pagination
    more_pages = True
    while more_pages:
        # Wait for listings to load
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".wrapper-main.mb-3")))
        except TimeoutException:
            print("Timeout waiting for listings to load.")
            break

        # Scrape the current page
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        # Update the following selector to match the actual HTML for a single listing
        listings = soup.find_all('div', class_='single-listing-selector')  

        for listing in listings:
            # Extract the lot number and the name from the listing
            lot_number, lot_name = get_lot_details_from_listing(listing)
            excel_lot = f"Lot - {lot_number}"
            
            # Check if lot number and name match
            if excel_lot in lot_to_name and lot_to_name[excel_lot] == lot_name:
                price = get_price_from_listing(listing)
                if price is not None:
                    # Find the corresponding row in Excel
                    for row_num in range(10, sheet.max_row + 1):
                        if sheet.cell(row=row_num, column=11).value == excel_lot:
                            sheet.cell(row=row_num, column=4).value = price  # Write the price to column D
                            break

        # Save after each page
        workbook.save(excel_path)

        # Check for a next page button and click it if present
        more_pages = click_next_page_button(driver)

    workbook.close()


def click_next_page_button(driver):
    try:
        next_page_button = driver.find_element(By.LINK_TEXT, "Next")
        if next_page_button.is_enabled():
            driver.execute_script("arguments[0].click();", next_page_button)
            WebDriverWait(driver, 10).until(EC.staleness_of(next_page_button))
            return True
    except Exception as e:
        print("No more pages or an error occurred:", e)
    return False


def main():
    driver = setup_selenium_driver()
    lot_to_name = read_lot_and_names_from_excel(EXCEL_PATH, SHEET_NAME)
    driver.get(AUCTION_URL)
    scrape_and_write_prices(driver, EXCEL_PATH, SHEET_NAME, lot_to_name)
    driver.quit()

if __name__ == "__main__":
    main()
